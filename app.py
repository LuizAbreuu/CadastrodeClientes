import openpyxl
import sqlite3
from tkinter import Tk, Label, Entry, Button, messagebox

# Função para adicionar dados à planilha
def adicionar_dados_planilha(nome, rg, cpf, local_embarque):
    # Abre o arquivo Excel
    planilha = openpyxl.load_workbook('dados.xlsx')
    
    # Seleciona a primeira planilha
    sheet = planilha.active
    
    # Adiciona os dados à próxima linha vazia
    linha = sheet.max_row + 1
    sheet.cell(row=linha, column=1).value = nome
    sheet.cell(row=linha, column=2).value = rg
    sheet.cell(row=linha, column=3).value = cpf
    sheet.cell(row=linha, column=4).value = local_embarque
    
    # Salva as alterações
    planilha.save('dados.xlsx')

# Função para salvar dados no banco de dados
def salvar_dados_banco(nome, rg, cpf, local_embarque):
    conn = sqlite3.connect('clientes.db')
    c = conn.cursor()
    
    # Insere os dados na tabela
    c.execute("INSERT INTO clientes (nome, rg, cpf, local_embarque) VALUES (?, ?, ?, ?)",
              (nome, rg, cpf, local_embarque))
    
    conn.commit()
    conn.close()

# Função para pesquisar dados no banco de dados por CPF
def pesquisar_por_cpf(cpf):
    conn = sqlite3.connect('clientes.db')
    c = conn.cursor()
    
    # Executa a consulta no banco de dados
    c.execute("SELECT * FROM clientes WHERE cpf=?", (cpf,))
    
    # Recupera os resultados
    resultados = c.fetchall()
    
    conn.close()
    
    return resultados

# Função para lidar com o evento do botão "Salvar"
def salvar():
    nome = entry_nome.get()
    rg = entry_rg.get()
    cpf = entry_cpf.get()
    local_embarque = entry_local.get()
    
    adicionar_dados_planilha(nome, rg, cpf, local_embarque)
    salvar_dados_banco(nome, rg, cpf, local_embarque)
    
    messagebox.showinfo('Sucesso', 'Os dados foram salvos!')

    limpar_campos()  # Limpa os campos após salvar

# Função para limpar os campos de entrada de dados
def limpar_campos():
    entry_nome.delete(0, 'end')
    entry_rg.delete(0, 'end')
    entry_cpf.delete(0, 'end')
    entry_local.delete(0, 'end')

# Função para lidar com o evento do botão "Pesquisar CPF"
def pesquisar():
    cpf = entry_cpf.get()
    
    resultados = pesquisar_por_cpf(cpf)
    
    if len(resultados) > 0:
        messagebox.showinfo('Resultados', f"Nome: {resultados[0][0]}\nRG: {resultados[0][1]}\nCPF: {resultados[0][2]}\nLocal de Embarque: {resultados[0][3]}")
    else:
        messagebox.showinfo('Resultados', 'Nenhum resultado encontrado.')

# Configuração da interface gráfica usando tkinter
root = Tk()
root.title('Cadastro de Clientes')
root.geometry('300x300')

label_nome = Label(root, text="Nome completo:")
label_nome.pack()
entry_nome = Entry(root)
entry_nome.pack()

label_rg = Label(root, text="RG:")
label_rg.pack()
entry_rg = Entry(root)
entry_rg.pack()

label_cpf = Label(root, text="CPF:")
label_cpf.pack()
entry_cpf = Entry(root)
entry_cpf.pack()

label_local = Label(root, text="Local de Embarque:")
label_local.pack()
entry_local = Entry(root)
entry_local.pack()

btn_salvar = Button(root, text="Salvar", command=salvar)
btn_salvar.pack()

btn_pesquisar = Button(root, text="Pesquisar CPF", command=pesquisar)
btn_pesquisar.pack()

btn_limpar = Button(root, text="Limpar", command=limpar_campos)
btn_limpar.pack()

root.mainloop()
